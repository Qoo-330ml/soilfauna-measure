"""Export CSV/XLSX and image artifacts tests."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest
from PIL import Image

from soilfauna_measure.core.calibration import build_scale_calibration
from soilfauna_measure.core.mask_operations import polygon_to_mask
from soilfauna_measure.exporters.table_export import (
    build_measurement_rows,
    export_csv,
    export_xlsx,
)
from soilfauna_measure.services.export_service import ExportOptions, export_project
from soilfauna_measure.services.object_service import ObjectService
from soilfauna_measure.storage.workspace import open_workspace, save_workspace


def _prep_workspace(tmp_path: Path) -> Path:
    root = tmp_path / "expws"
    root.mkdir()
    rgb = np.full((60, 80, 3), 240, dtype=np.uint8)
    Image.fromarray(rgb).save(root / "s.tif", format="TIFF")
    ws = open_workspace(root)
    rec = ws.images[0]
    rec.width, rec.height = 80, 60
    rec.scale = build_scale_calibration((0, 0), (80, 0), 800, "um")
    svc = ObjectService()
    svc.set_workspace(ws.root)
    svc.bind_image(rec, width=80, height=60)
    obj = svc.create_from_mask(
        rec,
        polygon_to_mask([[10, 10], [40, 10], [40, 40], [10, 40]], 60, 80),
        rec.scale,
        category_id="insect",
        confirmed=True,
    )
    svc.apply_length_points(obj, [[12, 12], [38, 38]], rec.scale)
    save_workspace(ws)
    return root


def test_build_rows_and_csv_xlsx(tmp_path: Path):
    root = _prep_workspace(tmp_path)
    ws = open_workspace(root)
    rows = build_measurement_rows(ws.project)
    assert len(rows) == 1
    assert rows[0]["object_id"].startswith("s_")
    assert rows[0]["category_id"] == "insect"
    assert rows[0]["category_zh"] == "昆虫"
    assert rows[0]["area_px"] > 0
    assert rows[0]["length_px"] != ""

    csv_path = export_csv(ws.project, tmp_path / "m.csv")
    assert csv_path.is_file()
    text = csv_path.read_text(encoding="utf-8-sig")
    assert "object_id" in text
    assert "insect" in text

    xlsx_path = export_xlsx(ws.project, tmp_path / "m.xlsx")
    assert xlsx_path.is_file()
    # open with openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws_sheet = wb.active
    headers = [c.value for c in next(ws_sheet.iter_rows(min_row=1, max_row=1))]
    assert "object_id" in headers
    assert ws_sheet.max_row == 2  # header + 1 data


def test_full_export_service(tmp_path: Path):
    root = _prep_workspace(tmp_path)
    ws = open_workspace(root)
    result = export_project(
        ws,
        ExportOptions(
            csv=True, xlsx=True, masks=True, crops=True, annotated=True
        ),
    )
    assert result.output_dir.is_dir()
    assert any(p.name == "measurements.csv" for p in result.files)
    assert any(p.name == "measurements.xlsx" for p in result.files)
    assert any("annotated" in str(p) for p in result.files)
    assert any(p.suffix == ".png" and "crop" in p.name for p in result.files)
    assert not result.errors
